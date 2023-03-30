import openpyxl
from datetime import datetime, date
import datetime
from dateutil import parser
import pytz
import openai
import os
from dotenv import load_dotenv

class DateInputExcel():
    
    def __init__(self, select_date, study_content, meeting_check, impressions, lessen_check, report_direction, ai_check):
        
        iso_conversion = parser.isoparse(select_date).replace(tzinfo=None)
        self.conversion = iso_conversion
        self.test_date_str = self.conversion.strftime('%Y-%m-%d-%a')
        self.study_content = study_content
        self.impressions = impressions
        self.report_direction = report_direction
        self.ai_check = ai_check
        
        if meeting_check == "yes":
            self.meeting_content = "・社内ミーティング"
        else:
            self.meeting_content = ""
        if lessen_check == "yes":
            self.lessen_content = "・プログラミングレッスン"
        else:
            self.lessen_content = ""
            
            
        # AIによる自動文章作成機能を使用し、所感にデータを入れる。
        if ai_check == "yes":

            load_dotenv()

            # APIキーを取得
            openai.api_key = os.environ["OPENAI_API_KEY"]

            # モデルを指定
            model_engine = "text-davinci-003"

            # 指示（Prompt）を設定
            prompt = "私の代わりに日報の所感を書いてください。文字数は150文字で、今日実施した内容は"+self.study_content +"です。"\
                "所感の方向性としては"+ self.report_direction + "の感じでなるべく自然な文章になるように作成してください。"

            # 推論を実行
            completions = openai.Completion.create(
                engine=model_engine,
                prompt=prompt,
                max_tokens=1024,
                n=1,
                stop=None,
                temperature=0.5,
            )

            # 推論結果を出力
            message = completions.choices[0].text
            self.impressions = message

    def return_impression(self):
        return self.impressions

    def date_input(self):

        test_path = "日報.xlsx"
        wb = openpyxl.load_workbook(test_path, data_only=True)
        last_sheet = wb.worksheets[-1]

        # 作成したい日報の曜日日付(週)を作成する
        last_sheet["H11"].value = last_sheet["H10"].value + datetime.timedelta(days=1)
        last_sheet["H12"].value = last_sheet["H10"].value + datetime.timedelta(days=2)
        last_sheet["H13"].value = last_sheet["H10"].value + datetime.timedelta(days=3)
        last_sheet["H14"].value = last_sheet["H10"].value + datetime.timedelta(days=4)
        last_sheet["H15"].value = last_sheet["H10"].value + datetime.timedelta(days=5)
        last_sheet["H16"].value = last_sheet["H10"].value + datetime.timedelta(days=6)

        # 作成したい日報の曜日に基づき、「月日」欄の「月」を入力する
        last_sheet["A10"].value = int(str(last_sheet["H10"].value)[6])
        last_sheet["A16"].value = int(str(last_sheet["H11"].value)[6])
        last_sheet["A22"].value = int(str(last_sheet["H12"].value)[6])
        last_sheet["A28"].value = int(str(last_sheet["H13"].value)[6])
        last_sheet["A34"].value = int(str(last_sheet["H14"].value)[6])
        last_sheet["A40"].value = int(str(last_sheet["H15"].value)[6])
        last_sheet["A46"].value = int(str(last_sheet["H16"].value)[6])

        # 作成したい日報の曜日に基づき、「月日」欄の「日」を入力する
        last_sheet["A11"].value = int(str(last_sheet["H10"].value)[8:10])
        last_sheet["A17"].value = int(str(last_sheet["H11"].value)[8:10])
        last_sheet["A23"].value = int(str(last_sheet["H12"].value)[8:10])
        last_sheet["A29"].value = int(str(last_sheet["H13"].value)[8:10])
        last_sheet["A35"].value = int(str(last_sheet["H14"].value)[8:10])
        last_sheet["A41"].value = int(str(last_sheet["H15"].value)[8:10])
        last_sheet["A47"].value = int(str(last_sheet["H16"].value)[8:10])
        last_sheet.title = str(last_sheet["H10"].value)[0:4] + str(last_sheet["H10"].value)[5:7] + str(last_sheet["H10"].value)[8:10] + "_" + str(last_sheet["H16"].value)[0:4] + str(last_sheet["H16"].value)[5:7] + str(last_sheet["H16"].value)[8:10]
        wb.save("日報.xlsx")

    def sell_seach(self):
        
        test_path = "日報.xlsx"
        wb = openpyxl.load_workbook(test_path, data_only=True)
        
        for date_check in wb.worksheets:

            if str(date_check["H10"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C9"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                break

            if str(date_check["H11"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C15"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                break

            if str(date_check["H12"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C21"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                break

            if str(date_check["H13"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C27"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                break

            if str(date_check["H14"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C33"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                break

    def Input_excel(self):
        
        test_path = "日報.xlsx"
        wb = openpyxl.load_workbook(test_path, data_only=True)


        for date_check in wb.worksheets:

            if str(date_check["H10"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C9"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                return True

            elif str(date_check["H11"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C15"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                return True

            elif str(date_check["H12"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C21"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                return True

            elif str(date_check["H13"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C27"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                return True

            elif str(date_check["H14"].value)[0:10] == str(self.test_date_str)[0:10]:
                sell_num = "C33"
                self.create_impressions(date_check, sell_num)
                wb.save("日報.xlsx")
                return True
        return False

    def value_set(self):

        # WBを開き、最初のシートをコピーする
        test_path = "日報.xlsx"
        wb = openpyxl.load_workbook(test_path, data_only=True)
        first_sheet = wb.worksheets[0]
        wb.copy_worksheet(first_sheet)
        wb.save("日報.xlsx")
        # 作成したい日報の曜日を算出する
        week_day = self.test_date_str[11:14]
        # コピーしたシートを開く(一番後ろのシートを開く)
        last_sheet = wb.worksheets[-1]

        if week_day == "Fri":
            date_logic = 4
            last_sheet["H10"].value = self.conversion - datetime.timedelta(days=date_logic)
            wb.save("日報.xlsx")
            self.date_input()
            self.sell_seach()

        if week_day == "Thu":
            date_logic = 3
            last_sheet["H10"].value = self.conversion - datetime.timedelta(days=date_logic)
            wb.save("日報.xlsx")
            self.date_input()
            self.sell_seach()

        if week_day == "Wed":
            date_logic = 2
            last_sheet["H10"].value = self.conversion - datetime.timedelta(days=date_logic)
            wb.save("日報.xlsx")
            self.date_input()
            self.sell_seach()

        if week_day == "Tue":
            date_logic = 1
            last_sheet["H10"].value = self.conversion - datetime.timedelta(days=date_logic)
            wb.save("日報.xlsx")
            self.date_input()
            self.sell_seach()

        if week_day == "Sat":
            raise ValueError("土曜日の日報は作成できません")

        if week_day == "Sun":
            raise ValueError("日曜日の日報は作成できません")

        if week_day == "Mon":

            last_sheet["H10"].value = self.conversion
            wb.save("日報.xlsx")
            self.date_input()
            self.sell_seach()

    def create_impressions(self, date_check, sell_num):
        impresson = f"【プログラミング研修業務】\n・{self.study_content}  {self.meeting_content}  {self.lessen_content} ・SNS広報作業\n【所感】{self.impressions}"
        date_check[sell_num].value = impresson